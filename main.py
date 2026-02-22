import pandas as pd
import asyncio
from playwright.async_api import async_playwright
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import sys
import os

# Configurações
SITE_URL = "https://www.siac.pt/pt"
INPUT_SELECTOR = "input[placeholder*='num. transponder']" # Ajustado com base na imagem
RESULT_SUCCESS_TEXT = "Animal com registo no SIAC"

async def check_siac(context, microchip):
    page = await context.new_page()
    try:
        await page.goto(SITE_URL)
        # Esperar pelo input. O seletor pode precisar de ajuste fino se o placeholder mudar ligeiramente
        try:
            input_field = await page.wait_for_selector("input", timeout=10000)
            # Tentar encontrar o input correto se houver vários
            inputs = await page.query_selector_all("input")
            target_input = None
            for inp in inputs:
                placeholder = await inp.get_attribute("placeholder")
                if placeholder and "transponder" in placeholder.lower():
                    target_input = inp
                    break
            
            if not target_input:
                target_input = input_field

            await target_input.fill("")
            await target_input.type(str(microchip), delay=100)
            
            # O site parece validar automaticamente ou após o enter/perda de foco
            # Vamos aguardar um pouco pelo resultado dinâmico
            await asyncio.sleep(2)
            
            # Capturar o texto que aparece após a inserção
            # Na imagem, o resultado aparece logo abaixo ou num popup interno
            content = await page.content()
            if RESULT_SUCCESS_TEXT in content:
                return "Válido (Registado no SIAC)"
            else:
                return "Não Registado ou Inválido"
        except Exception as e:
            return f"Erro na consulta: {str(e)}"
    finally:
        await page.close()

async def main(excel_file):
    if not os.path.exists(excel_file):
        print(f"Erro: Arquivo {excel_file} não encontrado.")
        return

    print(f"Lendo arquivo: {excel_file}")
    df = pd.read_excel(excel_file)
    
    # Supõe que a primeira coluna tem os números de microchip
    # Pode ser ajustado se o utilizador indicar outra coluna
    column_name = df.columns[0]
    microchips = df[column_name].tolist()
    
    results = []
    
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context()
        
        print(f"A validar {len(microchips)} números...")
        for i, chip in enumerate(microchips):
            if pd.isna(chip):
                results.append("N/A")
                continue
            
            print(f"[{i+1}/{len(microchips)}] Validando: {chip}")
            res = await check_siac(context, str(int(chip)) if isinstance(chip, float) else str(chip))
            results.append(res)
        
        await browser.close()

    # Escrever resultados de volta no Excel com formatação
    df['Resultado SIAC'] = results
    output_file = excel_file # Sobreescreve o original conforme solicitado
    df.to_excel(output_file, index=False)
    
    # Aplicar cores com openpyxl
    wb = load_workbook(output_file)
    ws = wb.active
    
    # Encontrar a coluna do 'Resultado SIAC' (é a última agora)
    res_col_idx = len(df.columns)
    
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=res_col_idx)
        if cell.value and "Válido" in str(cell.value):
            cell.fill = green_fill
        elif cell.value and ("Não Registado" in str(cell.value) or "Erro" in str(cell.value)):
            cell.fill = red_fill
            
    wb.save(output_file)
    print(f"Concluído! Resultados guardados em: {output_file}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python main.py <nome_do_arquivo.xlsx>")
    else:
        asyncio.run(main(sys.argv[1]))
